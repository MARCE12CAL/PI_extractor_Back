"""
==============================================================================
CONTROLADOR DE EXPORTACIÓN - GENERACIÓN DE EXCEL CONSOLIDADO
==============================================================================

Descripción:
    Controlador Singleton que gestiona la exportación de datos extraídos
    a formato Excel estructurado. Consolida todos los campos detectados
    en múltiples documentos en una única tabla lista para migración a BD.

Características:
    - Patrón Singleton para gestión única de recursos
    - Consolidación de datos de múltiples documentos
    - Generación de Excel con formato tabular
    - Columnas dinámicas según campos detectados
    - Limpieza y normalización de datos

Autor: EMPROSERVIS
Versión: 1.0.0
==============================================================================
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models.scanned_document import ScannedDocument
from models.document_field import DocumentField
from models.scan_job import ScanJob


class ExcelExporter:
    """
    Controlador de exportación a Excel con patrón Singleton.
    
    Gestiona la consolidación de campos extraídos de múltiples documentos
    y su exportación a formato Excel estructurado.
    """
    
    _instance = None
    
    def __new__(cls):
        """
        Implementación del patrón Singleton.
        Garantiza una única instancia del controlador.
        """
        if cls._instance is None:
            cls._instance = super(ExcelExporter, cls).__new__(cls)
        return cls._instance
    
    def export_to_excel(self, job_id):
        """
        Exporta todos los campos de un trabajo a Excel estructurado.
        
        Consolida los campos extraídos de todos los documentos de un
        trabajo en una tabla Excel con columnas organizadas.
        
        Args:
            job_id (int): ID del trabajo de escaneo
            
        Returns:
            str: Ruta del archivo Excel generado
            
        Raises:
            Exception: Si no hay documentos o error durante generación
        """
        try:
            # Obtener el trabajo
            trabajo = ScanJob.query.get(job_id)
            if not trabajo:
                raise Exception(f"Trabajo {job_id} no encontrado")
            
            # Obtener todos los documentos del trabajo
            documentos = ScannedDocument.query.filter_by(scan_job_id=job_id).all()
            
            if not documentos:
                raise Exception("No hay documentos para exportar")
            
            # Extraer todos los campos únicos detectados
            campos_unicos = self._obtener_campos_unicos(documentos)
            
            # Construir datos en formato tabular
            datos_tabla = self._construir_tabla_datos(documentos, campos_unicos)
            
            # Crear DataFrame de pandas
            df = pd.DataFrame(datos_tabla)
            
            # Generar Excel
            ruta_excel = self._generar_excel(df, job_id, trabajo)
            
            return ruta_excel
        
        except Exception as error:
            raise Exception(f"Error al exportar a Excel: {str(error)}")
    
    def _obtener_campos_unicos(self, documentos):
        """
        Obtiene todos los campos únicos detectados en los documentos.
        
        Args:
            documentos (list): Lista de ScannedDocument
            
        Returns:
            list: Lista de nombres de campos únicos ordenados
        """
        campos_set = set()
        
        for documento in documentos:
            for campo in documento.fields:
                campos_set.add(campo.field_name.lower())
        
        # Ordenar campos con prioridad
        campos_prioritarios = ['nombre', 'apellido', 'cedula', 'dni', 'fecha']
        campos_ordenados = []
        
        # Agregar campos prioritarios primero
        for campo in campos_prioritarios:
            if campo in campos_set:
                campos_ordenados.append(campo)
                campos_set.remove(campo)
        
        # Agregar resto de campos ordenados alfabéticamente
        campos_ordenados.extend(sorted(list(campos_set)))
        
        return campos_ordenados
    
    def _construir_tabla_datos(self, documentos, campos_unicos):
        """
        Construye una tabla de datos consolidada.
        
        Args:
            documentos (list): Lista de ScannedDocument
            campos_unicos (list): Lista de nombres de campos
            
        Returns:
            list: Lista de diccionarios con datos de cada documento
        """
        datos_tabla = []
        
        for documento in documentos:
            # Crear fila para este documento
            fila = {}
            
            # Inicializar todas las columnas con valores vacíos
            for campo_nombre in campos_unicos:
                fila[campo_nombre] = ''
            
            # Llenar con valores detectados
            for campo in documento.fields:
                nombre_campo = campo.field_name.lower()
                if nombre_campo in campos_unicos:
                    # Limpiar y normalizar valor
                    valor = self._limpiar_valor(campo.field_value)
                    fila[nombre_campo] = valor
            
            # Agregar metadatos
            fila['_archivo_origen'] = Path(documento.file_path).name
            fila['_confianza'] = f"{documento.confidence_score:.1f}%"
            fila['_fecha_procesamiento'] = documento.scanned_at.strftime('%Y-%m-%d %H:%M')
            
            datos_tabla.append(fila)
        
        return datos_tabla
    
    def _limpiar_valor(self, valor):
        """
        Limpia y normaliza un valor extraído.
        
        Args:
            valor (str): Valor a limpiar
            
        Returns:
            str: Valor limpio
        """
        if not valor:
            return ''
        
        # Eliminar espacios extras
        valor = ' '.join(valor.split())
        
        # Capitalizar nombres propios
        if len(valor) > 0 and valor[0].isalpha():
            valor = valor.title()
        
        return valor
    
    def _generar_excel(self, df, job_id, trabajo):
        """
        Genera archivo Excel con formato profesional.
        
        Args:
            df (DataFrame): Datos a exportar
            job_id (int): ID del trabajo
            trabajo (ScanJob): Modelo del trabajo
            
        Returns:
            str: Ruta del archivo Excel generado
        """
        # Crear directorio de salida
        ruta_salida = Path('outputs') / f'datos_consolidados_job_{job_id}.xlsx'
        ruta_salida.parent.mkdir(parents=True, exist_ok=True)
        
        # Exportar DataFrame a Excel
        with pd.ExcelWriter(str(ruta_salida), engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos Extraídos', index=False)
        
        # Aplicar formato profesional
        self._aplicar_formato_excel(str(ruta_salida), trabajo)
        
        return str(ruta_salida)
    
    def _aplicar_formato_excel(self, ruta_excel, trabajo):
        """
        Aplica formato profesional al archivo Excel.
        
        Args:
            ruta_excel (str): Ruta del archivo Excel
            trabajo (ScanJob): Modelo del trabajo
        """
        try:
            wb = load_workbook(ruta_excel)
            ws = wb['Datos Extraídos']
            
            # Estilo de encabezado
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            border_style = Side(style='thin', color='000000')
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            
            # Aplicar estilo a encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar bordes a todas las celdas con datos
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
            
            # Agregar hoja de metadatos
            ws_meta = wb.create_sheet('Información del Trabajo')
            
            metadatos = [
                ['Información del Trabajo de Escaneo', ''],
                ['', ''],
                ['ID del Trabajo:', trabajo.id],
                ['Carpeta Procesada:', trabajo.folder_path],
                ['Estado:', trabajo.status],
                ['Total de Archivos:', trabajo.total_files],
                ['Archivos Procesados:', trabajo.processed_files],
                ['Fecha de Creación:', trabajo.created_at.strftime('%Y-%m-%d %H:%M:%S')],
                ['Fecha de Finalización:', trabajo.completed_at.strftime('%Y-%m-%d %H:%M:%S') if trabajo.completed_at else 'En proceso'],
                ['', ''],
                ['Generado por:', 'Sistema OCR EMPROSERVIS'],
                ['Fecha de Exportación:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ]
            
            for row_idx, row_data in enumerate(metadatos, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_meta.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1 and value:
                        cell.font = Font(bold=True)
            
            # Ajustar ancho de columnas en hoja de metadatos
            ws_meta.column_dimensions['A'].width = 30
            ws_meta.column_dimensions['B'].width = 50
            
            # Guardar cambios
            wb.save(ruta_excel)
            wb.close()
        
        except Exception as error:
            print(f"Error aplicando formato: {str(error)}")
    
    @classmethod
    def get_instance(cls):
        """
        Obtiene la instancia única del controlador (Singleton).
        
        Returns:
            ExcelExporter: Instancia única del controlador
        """
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance


"""
==============================================================================
DOCUMENTACIÓN TÉCNICA - EXCEL EXPORTER
==============================================================================

PROPÓSITO DEL CONTROLADOR
==========================

El ExcelExporter es responsable de:

1. Consolidación de Datos:
   - Extraer campos de múltiples documentos
   - Identificar columnas únicas
   - Normalizar valores
   - Organizar en formato tabular

2. Generación de Excel:
   - Crear archivo XLSX estructurado
   - Aplicar formato profesional
   - Incluir metadatos del trabajo
   - Optimizar para migración a BD


ESTRUCTURA DEL EXCEL GENERADO
==============================

Hoja 1: "Datos Extraídos"
--------------------------
| nombre  | cedula     | fecha      | _archivo_origen | _confianza | _fecha_procesamiento |
|---------|------------|------------|-----------------|------------|----------------------|
| Marcela | 2300434524 | 23/11/2025 | factura1.pdf    | 92.5%      | 2025-01-15 10:30    |
| Juan    | 1700234567 | 15/10/2025 | doc2.jpg        | 87.3%      | 2025-01-15 10:31    |

Características:
- Encabezados en azul con texto blanco
- Bordes en todas las celdas
- Ancho de columnas ajustado automáticamente
- Valores normalizados y limpios

Hoja 2: "Información del Trabajo"
----------------------------------
Metadatos del trabajo:
- ID del trabajo
- Carpeta procesada
- Estado del trabajo
- Estadísticas de procesamiento
- Fechas de creación y finalización


CAMPOS PRIORITARIOS
====================

Los siguientes campos aparecen primero en el Excel:
1. nombre
2. apellido
3. cedula
4. dni
5. fecha

Resto de campos ordenados alfabéticamente.


EJEMPLO DE USO
==============

from controllers.excel_exporter import ExcelExporter

# Exportar datos de un trabajo
exporter = ExcelExporter.get_instance()
ruta_excel = exporter.export_to_excel(job_id=123)

print(f"Excel generado: {ruta_excel}")
# Salida: outputs/datos_consolidados_job_123.xlsx


INTEGRACIÓN CON API
===================

Endpoint recomendado:
GET /api/documents/export-excel/<job_id>

from controllers.excel_exporter import ExcelExporter

@document_bp.route('/export-excel/<int:job_id>', methods=['GET'])
def export_to_excel(job_id):
    try:
        exporter = ExcelExporter.get_instance()
        ruta_excel = exporter.export_to_excel(job_id)
        
        return send_file(
            ruta_excel,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'datos_consolidados_{job_id}.xlsx'
        )
    except Exception as error:
        return jsonify({
            'success': False,
            'message': str(error)
        }), 500

==============================================================================
"""
